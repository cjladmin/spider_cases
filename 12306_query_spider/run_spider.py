# encoding: utf-8
# @Time : 2022/5/25 13:15
# @Author : Torres-圣君
# @File : run_spider.py
# @Sofaware : PyCharm
import requests
import json
from openpyxl import Workbook
from prettytable import PrettyTable
from save_city_list import get_city_data


class GetTrains:
    def __init__(self, date, begin_id, end_id):
        self.url = "https://kyfw.12306.cn/otn/leftTicket/query"
        # 构建请求头
        self.headers = {
            # 失效时，需要更新cookie
            "Cookie": "JSESSIONID=5BCD4997EB7387D6F2F26CF860144AE6; RAIL_EXPIRATION=1653658158853; RAIL_DEVICEID=OYdRuCkXuonxJIyWihWNwMa5x-JAFt30BYWuZd9lAzHOtXh1TezSjz0oQm9n0TYq3InM3pJKfGexQCQEFpOqkTJq5XqXQ_taNYf1hTlQ6YWdWKWrJosRmvmDdUmt9omgZ2sDBAmcohSg662SJ-55JM97DtJQ0sfA; guidesStatus=off; highContrastMode=defaltMode; cursorStatus=off; BIGipServerotn=384827914.50210.0000; BIGipServerpool_passport=31719946.50215.0000; route=c5c62a339e7744272a54643b3be5bf64; _jc_save_toDate=2022-05-25; _jc_save_wfdc_flag=dc; _jc_save_fromStation=%u5546%u4E18%2CSQF; _jc_save_toStation=%u90D1%u5DDE%2CZZF; _jc_save_fromDate=2022-05-26",
            # "Referer": referer,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.64 Safari/537.36 Edg/101.0.1210.53"
        }
        # 构建请求所需参数
        self.params = {
            "leftTicketDTO.train_date": date,
            "leftTicketDTO.from_station": begin_id,
            "leftTicketDTO.to_station": end_id,
            "purpose_codes": "ADULT"
        }
        # 实例化美化表格对象
        self.pt = PrettyTable()

    def run(self):
        # 对目标网址发送请求
        res = requests.get(self.url, headers=self.headers, params=self.params).json()
        data_list = res['data']['result']
        # 构造表格的表头，用于展示和保存
        header_list = [
            ['车次', '出发时间', '到达时间', '历时', '商务座', '一等座', '二等座', '软卧', '硬卧', '硬座', '无座', '备注']
        ]
        # 将表头信息添加进展示表格的表头
        self.pt.field_names = header_list[0]
        for data in data_list:
            # 格式化添加表数据
            trains_msg = self.format_data(data)
            # 将数据添加进列表，用于保存
            header_list.append(trains_msg)
        # 打印表格
        print(self.pt)
        # 返回车次信息列表
        return header_list

    def format_data(self, data):
        # 将返回的数据以'|'进行分隔
        all_data_list = data.split('|')
        # 提取车次的信息
        trains_msg = [
            all_data_list[3],
            all_data_list[8],
            all_data_list[9],
            all_data_list[10],
            all_data_list[32] if all_data_list[32] != "" else "--",
            all_data_list[31] if all_data_list[31] != "" else "--",
            all_data_list[30] if all_data_list[30] != "" else "--",
            all_data_list[23] if all_data_list[23] != "" else "--",
            all_data_list[28] if all_data_list[28] != "" else "--",
            all_data_list[29] if all_data_list[29] != "" else "--",
            all_data_list[26] if all_data_list[26] != "" else "--",
            all_data_list[1] if all_data_list[1] != "" else "--"
        ]
        # 增添表内容
        self.pt.add_row(trains_msg)
        # 将提取的信息返回，用于保存
        return trains_msg

    def save_data(self, trains_data_list, date, begin, end):
        num = input("如果展示不清晰，需要保存时请扣1：")
        if num == "1":
            wb = Workbook()
            sheet = wb.create_sheet("车次信息", -1)
            # 遍历表格索引，写入数据
            for x in range(len(trains_data_list)):
                for y in range(len(trains_data_list[x])):
                    sheet.cell(x + 1, y + 1).value = trains_data_list[x][y]
            wb.save(f"./data/{date}_{begin}_{end}.xlsx")
            print("数据保存完成！")


if __name__ == '__main__':
    # 更新城市对应的英文代码，需要时再启用
    # get_city_data()
    date = input("请输入出发日期(YYYY-MM-DD)：")
    begin = input("请输入出发地：")
    end = input("请输入目的地：")
    # 读取生成的json文件
    city_list = json.load(open('./data/city_data.json', 'r'))
    # 获取城市对应的英文代码
    begin_id = city_list[begin]
    end_id = city_list[end]
    gt = GetTrains(date, begin_id, end_id)
    trains_data_list = gt.run()
    # 是否需要保存数据
    gt.save_data(trains_data_list, date, begin, end)
    print(
        "12306直达链接(复制到浏览器打开)：",
        "https://kyfw.12306.cn/otn/leftTicket/init?"
        "linktypeid=dc&"
        f"fs={begin},{begin_id}&"
        f"ts={end},{end_id}&"
        f"date={date}&"
        "flag=N,N,Y"
    )
