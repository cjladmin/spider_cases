# encoding: utf-8
# @Time : 2022/6/14 16:16
# @Author : Torres-圣君
# @File : download_fonts.py
# @Sofaware : PyCharm
import re
import time
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from selenium import webdriver
import ua_ip_pool


class DoubanMovies:
    def __init__(self, links: list):
        options = webdriver.ChromeOptions()
        # 无头模式
        options.add_argument('--headless')
        options.add_argument('--disable-gpu')
        # 移除指纹
        options.add_experimental_option('excludeSwitches', ['enable-automation'])
        options.add_experimental_option('useAutomationExtension', False)
        self.driver = webdriver.Chrome(options=options)
        self.driver.execute_cdp_cmd(
            'Page.addScriptToEvaluateOnNewDocument',
            {
                'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'
            }
        )
        self.links = links
        self.headers = {
            "User-Agent": ua_ip_pool.get_ua()
        }
        self.proxies = ua_ip_pool.get_proxies()
        self.align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # 计数器
        self.count = 1

    def run(self):
        # 存放所有电影榜单信息
        for url in self.links:
            print("正在访问：", url)
            # 获取数据
            movie_data = self.get_move_links(url)
            # 展示数据
            print(movie_data)
            # 保存数据
            self.save_data(movie_data)
        # 关闭浏览器
        self.driver.close()

    def get_move_links(self, url):
        self.driver.get(url)
        # 等待页面加载动画
        time.sleep(10)
        # 返回加载后的源码
        page_source = self.driver.page_source
        # 提取热榜电影名称
        title_list = re.findall(r'title="(.*?)"', page_source)
        # 提取热榜电影链接
        url_list = re.findall(r'href="https://movie.douban.com/subject/(.*?)/', page_source)
        if len(url_list) == 0:
            url_list = re.findall(r'href="https://m.douban.com/movie/subject/(.*?)/', page_source)
        if ('2018' in url) or ('2017' in url):
            # 提取热榜电影名称
            special_title = re.findall('target="_blank">(.*?)</a>', page_source)
            # 获取前18个名称，后将特殊字段插入列表
            titles = title_list[:18]
            titles.insert(0, special_title[0])
            titles.insert(10, special_title[2])
            movie_title = titles
            # 不打乱去重
            new_url_list = sorted(set(url_list), key=url_list.index)
            movie_link = [f'https://movie.douban.com/subject/{i}' for i in new_url_list]
        elif '2016' in url:
            movie_title = title_list[:20]
            # 不打乱去重
            new_url_list = sorted(set(url_list), key=url_list.index)
            movie_link = [f'https://movie.douban.com/subject/{i}' for i in new_url_list]
        elif '2015' in url:
            movie_title = title_list[:20]
            # 不打乱去重
            new_url_list = sorted(set(url_list[:23]), key=url_list[:23].index)
            movie_link = [f'https://movie.douban.com/subject/{i}' for i in new_url_list]
        else:
            movie_title = title_list[:20]
            movie_link = [f'https://movie.douban.com/subject/{i}' for i in url_list[:20]]
        # 获取影片详细信息
        pingfen_list, yingchang_list, leixing_list, diqu_list = self.send_request(movie_link)
        # 返回获取的数据
        links_data = [
            movie_title, movie_link, pingfen_list, yingchang_list, leixing_list, diqu_list
        ]
        return links_data

    def send_request(self, movie_link):
        pingfen_list = []
        yingchang_list = []
        leixing_list = []
        diqu_list = []
        for url in movie_link:
            print(f"正在获取：'{url}'")
            while True:
                try:
                    res = requests.get(url, headers=self.headers, proxies=self.proxies, timeout=2)
                    if res.status_code == 200:
                        res = res.text
                        break
                except:
                    print("代理超时，正在更换代理！")
            # 评分
            pingfen_list.append(re.findall('property="v:average">(.*?)<.strong>', res)[0])
            # 影长
            yingchang_list.append(re.findall('property="v:runtime" content="(.*?)"', res)[0])
            # 类型
            leixing_list.append('/'.join(re.findall('property="v:genre">(.*?)</span>', res)))
            # 制片区
            diqu_list.append(re.findall('制片国家/地区:</span>(.*?)<br/>', res)[0].strip(' '))
        return pingfen_list, yingchang_list, leixing_list, diqu_list

    def save_data(self, movie_data):
        # 首次写入时，创建表格并添加表头
        if self.count == 1:
            # 创建新的excel表格
            wb = Workbook()
            sheet = wb.create_sheet("sheet1", -1)
            # 设置列宽
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 50
            sheet.column_dimensions['C'].width = 10
            sheet.column_dimensions['D'].width = 10
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 25
            excel_title = ['电影名', '电影链接', '电影评分', '电影影长', '电影类型', '制片地区']
            for x in range(len(excel_title)):
                sheet.cell(1, x+1).value = excel_title[x]
                # 居中对齐
                sheet.cell(1, 1).alignment = self.align
            self.count += 1
        # 后则读取并追加
        else:
            wb = load_workbook("榜单电影链接.xlsx")
            sheet = wb["sheet1"]
        for x in range(len(movie_data)):
            for y in range(len(movie_data[x])):
                sheet.cell(y + self.count, x + 1).value = movie_data[x][y]
                # 居中对齐
                sheet.cell(y + self.count, x + 1).alignment = self.align
        # 增加计数器
        self.count += 20
        # 保存该Excel表格
        wb.save("榜单电影链接.xlsx")


if __name__ == '__main__':
    dm = DoubanMovies(
        [f"https://movie.douban.com/annual/20{i}?source=navigation" for i in range(21, 14, -1)]
    )
    dm.run()
