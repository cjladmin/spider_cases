# encoding: utf-8
# @Time : 2022/6/23 9:53
# @Author : Torres-圣君
# @File : download_fonts.py
# @Software : PyCharm
import time
import requests
import json
from decryption_AES import AES_Decrypt
from get_message import OtherData
from openpyxl import Workbook
from openpyxl import load_workbook


class GetAESData:
    def __init__(self):
        self.url = 'http://gbcdn.dfcfw.com/rank/popularityList.js'
        self.headers = {
            "Referer": "http://guba.eastmoney.com/",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.124 Safari/537.36 Edg/102.0.1245.44"
        }
        self.gt = OtherData(self.headers)
        self.count = 1

    def run(self):
        types_list = ['A股市场', '港股市场', '美股市场']
        for types in range(0, 3):
            # 循环获取前五页top100
            for page in range(1, 6):
                time.sleep(1)
                print(f"正在获取第{page}页数据！")
                # 构建请求参数
                params = self.build_params(types, page)
                # 解密数据
                decrypt_data = self.get_response(params)
                # 获取页面数据
                page_all_data = self.format_data(decrypt_data)
                # 保存页面数据
                self.save_data(types_list[types], page_all_data)
                print(f"第{page}页数据保存完成！")
            # 计数器归1
            self.count = 1

    def build_params(self, types, page):
        """
        type: 0
        sort: 0
        page: 1
        v: 2022_6_23_9_56
        """
        t = time.localtime()
        time_list = time.strftime("%Y_%m_%d_%H_%M", t).split('_')
        now = '_'.join([i[-1] if i.startswith('0') else i for i in time_list])
        params = {
            "type": types,
            "sort": 0,
            "page": page,
            "v": now
        }
        print(params)
        return params

    def get_response(self, params):
        res = requests.get(self.url, headers=self.headers, params=params).text
        # 加密数据
        aes_data = res.split("'")[1]
        # 密钥
        key = 'ae13e0ad97cdd6e12408ac5063d88721'
        # 偏移量
        vi = 'getClassFromFile'
        # 使用AES解密
        decrypt_data = AES_Decrypt(key, vi, aes_data)
        return decrypt_data

    def format_data(self, decrypt_data):
        json_data = json.loads(decrypt_data)
        page_data = []
        for everyone in json_data:
            item = [
                everyone['rankNumber'],
                everyone['changeNumber'],
                everyone['code']
            ]
            page_data.append(item)
            print(item)
        page_all_data = self.gt.get_response(page_data)
        return page_all_data

    def save_data(self, title, page_all_data):
        # 首次保存需创建表格，并写入表头信息
        if self.count == 1:
            wb = Workbook()
            # 创建新的工作蒲
            sheet = wb.create_sheet('sheet1', -1)
            # 添加表头信息
            data_header = ['当前排名', '排名较昨日变动', '股票代码', '股票名称', '最新价', '涨跌额', '涨跌幅', '最高价', '最低价']
            page_all_data.insert(0, data_header)
        else:
            # 读取已有的工作蒲
            wb = load_workbook(f'./data/{title}_人气榜.xlsx')
            sheet = wb["sheet1"]
        for x in range(len(page_all_data)):
            for y in range(len(page_all_data[x])):
                sheet.cell(x + self.count, y + 1).value = page_all_data[x][y]
        # 保存表格并追加计数
        wb.save(f'./data/{title}_人气榜.xlsx')
        self.count += len(page_all_data)


if __name__ == '__main__':
    aes = GetAESData()
    aes.run()
