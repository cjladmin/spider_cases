# encoding: utf-8
# @Time : 2022/5/18 18:13
# @Author : Torres-圣君
# @File : get_types_user_msg.py
# @Sofaware : PyCharm
# https://www.huya.com/cache.php?m=LiveList&do=getLiveListByPage&gameId=1&tagAll=0&callback=getLiveListJsonpCallback&page=1
import requests
import json
import re
import threading
from openpyxl import Workbook
from get_proxyz import get_proxies
from get_ua import get_ua


class NowLiveUsers:
    def __init__(self, key, url, gameId):
        self.lock = threading.Lock()
        self.key = key
        self.url = "https://www.huya.com/cache.php"
        self.headers = {
            "Host": "www.huya.com",
            "Referer": url,
            "User-Agent": get_ua()
        }
        self.params = {
            "m": "LiveList",
            "do": "getLiveListByPage",
            "gameId": gameId,
            "tagAll": 0,
            "callback": "getLiveListJsonpCallback",
            "page": 1
        }

    def get_page_msg(self):
        # 创建一个用于汇总页面数据的列表
        all_users_data = [
            ["主播头像", "主播昵称", "房间ID号", "房间标题", "房间标签", "直播链接"]
        ]
        count = 0
        # 循环请求不同的页面
        while True:
            # 启用线程锁，防止数据穿线
            with self.lock:
                count += 1
                # 设置请求参数的页码值
                self.params['page'] = count
                # 对页面发送请求
                res = requests.get(self.url, headers=self.headers, params=self.params, proxies=get_proxies()).text
                # 使用re提取数据
                dict_data = re.findall(r'getLiveListJsonpCallback\((.*)\)', res)[0]
                json_data = json.loads(dict_data)
                data_list = json_data["data"]["datas"]
                # 如果页面返回为空时，跳出循环
                if len(data_list) != 0:
                    for data in data_list:
                        user_data = [
                            data['avatar180'],
                            data['nick'],
                            data['profileRoom'],
                            data['roomName'],
                            data['recommendTagName'],
                            "https://www.huya.com/" + data['profileRoom']
                        ]
                        # 将数据添加进页面汇总列表
                        all_users_data.append(user_data)
                    # 展示数据
                    print(all_users_data)
                else:
                    # 保存数据到Excel表格
                    self.save_data(all_users_data)
                    break

    def save_data(self, all_users_data_list):
        # 创建新的excel表格
        wb = Workbook()
        # 创建新的工作蒲
        sheet = wb.create_sheet(self.key, -1)
        # 遍历表格索引，写入数据
        for x in range(len(all_users_data_list)):
            for y in range(len(all_users_data_list[x])):
                sheet.cell(x+1, y+1).value = all_users_data_list[x][y]
        # 保存该文件
        wb.save(f"./data/{self.key}_直播用户信息.xlsx")
