# encoding: utf-8
# @Time : 2022/5/11 19:36
# @Author : Torres-圣君
# @File : save_data.py
# @Sofaware : PyCharm
from openpyxl import load_workbook


def save_data(data_list, title):
    # 创建Excel表对象
    wb = load_workbook("./data/猫眼实时数据.xlsx")
    # 创建新的sheet
    sheet = wb.create_sheet(title, -1)
    for i in range(0, len(data_list)):
        for j in range(0, len(data_list[i])):
            # 写入数据到单元格
            sheet.cell(row=i+1, column=j+1).value = data_list[i][j]
    # 保存并关闭文件
    wb.save("./data/猫眼实时数据.xlsx")
    print(f"{title}_写入成功...")
    wb.close()
