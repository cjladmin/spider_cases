# encoding: utf-8
# @Time : 2022/5/24 14:57
# @Author : Torres-圣君
# @File : run_spider.py
# @Software : PyCharm
import random
import time
import re
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains as ac
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment


class SaveTaobaoData:
    def __init__(self, search_content):
        # 搜索内容
        self.search_content = search_content
        # 数据计数器
        self.count = 1
        # 表格内容居中
        self.align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.options = webdriver.ChromeOptions()
        self.options.add_experimental_option('excludeSwitches', ['enable-automation'])
        self.options.add_experimental_option('useAutomationExtension', False)
        self.driver = webdriver.Chrome(options=self.options)
        self.driver.execute_cdp_cmd(
            'Page.addScriptToEvaluateOnNewDocument',
            {
                'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'
            }
        )

    def get_page(self):
        # 访问淘宝网址
        self.driver.get('https://www.taobao.com/')
        time.sleep(3)  # 停一会防止出意外
        # 向搜索框中添加内容，并按下回车进行搜索
        self.driver.find_element_by_xpath("//input[@aria-label='请输入搜索文字']").send_keys(self.search_content, Keys.ENTER)
        # 扫码登陆
        self.driver.find_element_by_xpath('//*[@id="login"]/div[1]/i').click()
        # 给20秒时间登陆自己的账号，根据自己的速度来
        time.sleep(20)
        # 进入循环获取每页数据信息
        self.get_next_page()

    def get_page_data(self):
        # 判断是否出现验证码
        self.driver = self.validation()
        # 模拟真人操作，拖动滚动条
        for x in range(1, 11, 2):
            time.sleep(0.5)
            j = x / 10
            js = 'document.documentElement.scrollTop = document.documentElement.scrollHeight * %f' % j
            self.driver.execute_script(js)
        # 页面存放的所有商品
        div_list = self.driver.find_elements_by_xpath('//*[@id="mainsrp-itemlist"]/div/div/div[1]/div')
        print("当前页面的总商品数：", len(div_list))
        # 首次数据添加表头
        if self.count == 1:
            data_list = [
                ['商品标题', '商品价格', '商品销量', '店铺名称', '商品链接']
            ]
        else:
            data_list = []
        for div in div_list:
            try:
                item = [
                    # 商品标题
                    div.find_element_by_xpath('./div[2]/div[2]/a').text.strip(" \t\n"),
                    # 商品价格
                    float(div.find_element_by_xpath('./div[2]/div[1]/div[1]/strong').text),
                    # 商品销量
                    div.find_element_by_xpath('./div[2]/div[1]/div[2]').text,
                    # 店铺名称
                    div.find_element_by_xpath('./div[2]/div[3]/div[1]/a/span[2]').text,
                    # 商品链接
                    div.find_element_by_xpath('./div/div/div[1]/a').get_attribute('href').strip(" \t\n")
                ]
                # 展示爬取到的数据
                print(item)
                # 追加进列表
                data_list.append(item)
            except:
                pass
        # 保存数据
        self.save_data(data_list)

    def get_next_page(self):
        # 判断是否出现验证码
        self.driver = self.validation()
        # 获取关键字商品的总页数
        get_page_number = self.driver.find_element_by_xpath('//*[@id="mainsrp-pager"]/div/div/div/div[1]').text
        page_number = int(re.findall(r'(\d+)', get_page_number)[0])
        print(f"共获取到数据：{page_number}页")
        # 循环访问所有页面
        for i in range(0, page_number*44, 44):
            # 构造每页的链接
            self.driver.get(f"https://s.taobao.com/search?q={self.search_content}&s={i}")
            # 隐式等待
            self.driver.implicitly_wait(10)
            # 解析页面数据
            self.get_page_data()
            print(f"第{int(i/44+1)}页数据写入完成！")

    def validation(self):
        content = self.driver.page_source
        if "亲，请拖动下方滑块完成验证" in content:
            con = self.hua_kuai()
            count = 1
            while "亲，请拖动下方滑块完成验证" in con and count <= 3:
                con = self.hua_kuai()
                count += 1
                if count == 3:
                    print("已尽力尝试自动滑动验证码，但抱歉没能通过，请手动滑一下吧~\n")
                    input("手动滑动后，请等待页面“加载完成”，扣1并按回车键继续采集：")
                    con = self.driver.page_source
        return self.driver

    def hua_kuai(self):
        ele = self.driver.find_element_by_xpath('//*[@id="nc_1_n1z"]')
        # 按住滑块元素不放
        ac(self.driver).click_and_hold(ele).perform()
        # 拖动滑块，xxx需要滑动的大小
        ac(self.driver).move_by_offset(300, random.randint(-5, 5)).perform()
        # 松开鼠标
        ac(self.driver).release().perform()
        # 加载页面
        time.sleep(2)
        try:
            # 点击重新滑动按钮
            self.driver.find_element_by_xpath('//*[@id="`nc_1_refresh1`"]').click()
        except:
            pass
        return self.driver.page_source

    def save_data(self, data_list):
        # 第一次写入需创建表格，后者追加内容
        if self.count == 1:
            # 创建新的excel表格
            wb = Workbook()
            sheet = wb.create_sheet("sheet1", -1)
            # 设置列宽
            sheet.column_dimensions['A'].width = 70
            sheet.column_dimensions['B'].width = 10
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 25
            sheet.column_dimensions['E'].width = 80
        else:
            wb = load_workbook(f"./data/{self.search_content}_商品信息.xlsx")
            sheet = wb["sheet1"]
        # 遍历表格索引，写入商品数据
        for x in range(len(data_list)):
            # 设置行高
            sheet.row_dimensions[x].height = 15
            for y in range(len(data_list[x])):
                sheet.cell(x + self.count, y + 1).value = data_list[x][y]
                # 居中显示
                sheet.cell(x + self.count, y + 1).alignment = self.align
        # 保存该Excel表格
        wb.save(f"./data/{self.search_content}_商品信息.xlsx")
        # 累加计数器，用于追加表格内容
        self.count += len(data_list)


if __name__ == '__main__':
    text = input("请输入需要搜索的关键字：")
    run_spider = SaveTaobaoData(text)
    run_spider.get_page()
